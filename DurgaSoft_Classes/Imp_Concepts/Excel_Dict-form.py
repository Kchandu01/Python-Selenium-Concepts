import openpyxl

class LoginPageData:

    #test_LoginPageData =
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        Dict1 = {}
        list1 = []
        book = openpyxl.load_workbook("D:\My Folder\TestData.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row+1): # to get rows
            #if sheet.cell(row = i, column= 1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):

                    Dict[sheet.cell(row = 1, column= j).value] = sheet.cell(row= i, column=j).value


                    list1.append(Dict)



                    print(list1)
                    print(len(list1))
        return [Dict]



    # for getting first column data( TestCase names)
    @staticmethod
    def getTestCaseName():
        li = []
        book = openpyxl.load_workbook("D:\My Folder\TestData.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row+1):
            li.append(sheet.cell(row=i, column=1).value)

        #print(li)
        return li


c = LoginPageData()

c.getTestData('TestCase2')
c.getTestCaseName()
